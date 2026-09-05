from datetime import datetime, date
from io import BytesIO
import re

from openpyxl import load_workbook

from django.db import transaction
from django.utils import timezone

from scoc_monitoring.models import (
    VoyageLeg,
    VoyageObservation,
)


# ============================================================
# HELPERS
# ============================================================

def clean_text(value):
    """
    Convert an Excel/message value to clean text.
    """

    if value is None:
        return ""

    return str(value).strip()


def to_float(value):
    """
    Safely convert a value to float.
    """

    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if not text:
        return None

    # Remove common characters
    text = text.replace(",", "")
    text = text.replace("MT", "")
    text = text.replace("mt", "")
    text = text.strip()

    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def parse_datetime(value):
    """
    Convert Excel/date/text value to a timezone-aware datetime.
    """

    if value is None:
        return None

    if isinstance(value, datetime):

        if timezone.is_naive(value):
            return timezone.make_aware(value)

        return value

    if isinstance(value, date):

        dt = datetime.combine(
            value,
            datetime.min.time(),
        )

        return timezone.make_aware(dt)

    text = str(value).strip()

    if not text:
        return None

    formats = [
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d-%b-%Y %H:%M",
        "%d-%b-%Y",
        "%d %b %Y %H:%M",
        "%d %b %Y",
    ]

    for fmt in formats:

        try:

            dt = datetime.strptime(
                text,
                fmt,
            )

            return timezone.make_aware(dt)

        except ValueError:
            continue

    return None


def normalize_vessel_name(name):
    """
    Normalize vessel names so that small formatting
    differences do not create duplicate voyage legs.
    """

    if not name:
        return ""

    text = str(name).strip().upper()

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text


# ============================================================
# VOYAGE LEG
# ============================================================

def get_or_create_voyage_leg(
    vessel_name="",
    voyage_reference="",
    departure="",
    destination="",
    voyage_route="",
    load_type="UNKNOWN",
    start_date=None,
    end_date=None,
    target_speed=None,
    target_consumption=None,
    source_message="",
):
    """
    Find an existing voyage leg or create a new one.
    """

    vessel_name = normalize_vessel_name(
        vessel_name
    )

    voyage_reference = clean_text(
        voyage_reference
    )

    departure = clean_text(
        departure
    )

    destination = clean_text(
        destination
    )

    voyage_route = clean_text(
        voyage_route
    )

    load_type = (
        load_type
        if load_type in {
            "LADEN",
            "BALLAST",
            "UNKNOWN",
        }
        else "UNKNOWN"
    )

    # --------------------------------------------------------
    # First try voyage reference
    # --------------------------------------------------------

    if voyage_reference:

        leg = VoyageLeg.objects.filter(
            vessel_name=vessel_name,
            voyage_reference=voyage_reference,
        ).order_by("-id").first()

        if leg:
            return leg, False

    # --------------------------------------------------------
    # Then try route + start date
    # --------------------------------------------------------

    if voyage_route and start_date:

        leg = VoyageLeg.objects.filter(
            vessel_name=vessel_name,
            voyage_route=voyage_route,
            start_date=start_date,
        ).first()

        if leg:
            return leg, False

    # --------------------------------------------------------
    # Create new leg
    # --------------------------------------------------------

    leg = VoyageLeg.objects.create(
        vessel_name=vessel_name,
        voyage_reference=voyage_reference,
        departure=departure,
        destination=destination,
        voyage_route=voyage_route,
        load_type=load_type,
        start_date=start_date,
        end_date=end_date,
        target_speed=target_speed,
        target_consumption=target_consumption,
        source_message=source_message,
    )

    return leg, True


# ============================================================
# UPDATE VOYAGE LEG
# ============================================================

def update_voyage_leg_summary(leg):
    """
    Recalculate summary values from observations.
    """

    observations = leg.observations.all()

    leg.observation_count = observations.count()

    # --------------------------------------------------------
    # Average speed
    # --------------------------------------------------------

    speeds = [
        obs.speed
        for obs in observations
        if obs.speed is not None
    ]

    if speeds:
        leg.average_speed = (
            sum(speeds) / len(speeds)
        )

    # --------------------------------------------------------
    # Average consumption
    # --------------------------------------------------------

    consumptions = [
        obs.consumption
        for obs in observations
        if obs.consumption is not None
    ]

    if consumptions:
        leg.average_consumption = (
            sum(consumptions)
            / len(consumptions)
        )

    # --------------------------------------------------------
    # Latest distance to go
    # --------------------------------------------------------

    latest = observations.order_by(
        "-reported_time",
        "-id",
    ).first()

    if latest:

        if latest.distance_to_go is not None:
            leg.distance_to_go = (
                latest.distance_to_go
            )

    # --------------------------------------------------------
    # End date
    # --------------------------------------------------------

    if latest:
        leg.end_date = latest.reported_time

    leg.save(
        update_fields=[
            "average_speed",
            "average_consumption",
            "distance_to_go",
            "observation_count",
            "end_date",
            "updated_at",
        ]
    )

    return leg


# ============================================================
# IMPORT ONE OBSERVATION
# ============================================================

@transaction.atomic
def save_observation(
    leg,
    data,
):
    """
    Save one parsed observation.

    `data` should be a dictionary whose keys match
    VoyageObservation fields.
    """

    reported_time = data.get(
        "reported_time"
    )

    if not reported_time:
        raise ValueError(
            "Observation has no reported_time."
        )

    # --------------------------------------------------------
    # Never duplicate same leg + reported time
    # --------------------------------------------------------

    observation, created = (
        VoyageObservation.objects.update_or_create(
            leg=leg,
            reported_time=reported_time,
            defaults={
                "vessel_name": data.get(
                    "vessel_name",
                    leg.vessel_name,
                ),

                "position": data.get(
                    "position",
                    "",
                ),

                "course": data.get(
                    "course"
                ),

                "speed": data.get(
                    "speed"
                ),

                "consumption": data.get(
                    "consumption"
                ),

                "distance": data.get(
                    "distance"
                ),

                "distance_run": data.get(
                    "distance_run"
                ),

                "distance_to_go": data.get(
                    "distance_to_go"
                ),

                "duration_days": data.get(
                    "duration_days"
                ),

                "running_hours": data.get(
                    "running_hours"
                ),

                "rpm": data.get(
                    "rpm"
                ),

                "power_kw": data.get(
                    "power_kw"
                ),

                "engine_power_kw": data.get(
                    "engine_power_kw"
                ),

                "shaft_power_kw": data.get(
                    "shaft_power_kw"
                ),

                "load_percent": data.get(
                    "load_percent"
                ),

                "engine_load_percent": data.get(
                    "engine_load_percent"
                ),

                "slip": data.get(
                    "slip"
                ),

                "hsfo_consumption_mt": data.get(
                    "hsfo_consumption_mt"
                ),

                "hsfo_rob": data.get(
                    "hsfo_rob"
                ),

                "hsfo_rob_mt": data.get(
                    "hsfo_rob_mt"
                ),

                "lsfo_consumption_mt": data.get(
                    "lsfo_consumption_mt"
                ),

                "lsfo_rob": data.get(
                    "lsfo_rob"
                ),

                "lsmgo_consumption_mt": data.get(
                    "lsmgo_consumption_mt"
                ),

                "lsmgo_rob": data.get(
                    "lsmgo_rob"
                ),

                "lsmgo_rob_mt": data.get(
                    "lsmgo_rob_mt"
                ),

                "cylinder_oil_consumption_l": data.get(
                    "cylinder_oil_consumption_l"
                ),

                "scoc": data.get(
                    "scoc"
                ),

                "wind": data.get(
                    "wind",
                    "",
                ),

                "swell": data.get(
                    "swell",
                    "",
                ),

                "current": data.get(
                    "current",
                    "",
                ),

                "eta": data.get(
                    "eta",
                    "",
                ),

                "remarks": data.get(
                    "remarks",
                    "",
                ),

                "source_file": data.get(
                    "source_file",
                    "",
                ),

                "source_message": data.get(
                    "source_message",
                    "",
                ),
            },
        )
    )

    return observation, created


# ============================================================
# EXCEL IMPORT
# ============================================================

@transaction.atomic
def import_excel(
    uploaded_file,
):
    """
    Main Excel import entry point.

    `uploaded_file` can be a Django UploadedFile.
    """

    workbook = load_workbook(
        uploaded_file,
        data_only=True,
    )

    imported = 0
    updated = 0

    # --------------------------------------------------------
    # Process each worksheet
    # --------------------------------------------------------

    for worksheet in workbook.worksheets:

        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            continue

        # ----------------------------------------------------
        # First row = headers
        # ----------------------------------------------------

        headers = [
            clean_text(value).lower()
            for value in rows[0]
        ]

        for row in rows[1:]:

            if not any(
                value is not None
                for value in row
            ):
                continue

            data = {}

            for index, value in enumerate(row):

                if index >= len(headers):
                    continue

                header = headers[index]

                if not header:
                    continue

                data[header] = value

            # ------------------------------------------------
            # Basic fields
            # ------------------------------------------------

            vessel_name = (
                data.get("vessel_name")
                or data.get("vessel")
                or data.get("ship")
                or ""
            )

            vessel_name = normalize_vessel_name(
                vessel_name
            )

            reported_time = (
                parse_datetime(
                    data.get("reported_time")
                )
                or parse_datetime(
                    data.get("date")
                )
                or parse_datetime(
                    data.get("report_date")
                )
            )

            if not reported_time:
                continue

            # ------------------------------------------------
            # Voyage information
            # ------------------------------------------------

            voyage_reference = clean_text(
                data.get(
                    "voyage_reference"
                )
                or data.get(
                    "voyage"
                )
            )

            departure = clean_text(
                data.get(
                    "departure"
                )
                or data.get(
                    "from"
                )
            )

            destination = clean_text(
                data.get(
                    "destination"
                )
                or data.get(
                    "to"
                )
            )

            voyage_route = clean_text(
                data.get(
                    "voyage_route"
                )
            )

            if (
                not voyage_route
                and (
                    departure
                    or destination
                )
            ):
                voyage_route = (
                    f"{departure} → "
                    f"{destination}"
                )

            # ------------------------------------------------
            # Load type
            # ------------------------------------------------

            load_type = clean_text(
                data.get(
                    "load_type"
                )
            ).upper()

            if load_type not in {
                "LADEN",
                "BALLAST",
            }:
                load_type = "UNKNOWN"

            # ------------------------------------------------
            # Get/create voyage leg
            # ------------------------------------------------

            leg, leg_created = (
                get_or_create_voyage_leg(
                    vessel_name=vessel_name,
                    voyage_reference=voyage_reference,
                    departure=departure,
                    destination=destination,
                    voyage_route=voyage_route,
                    load_type=load_type,
                    start_date=reported_time,
                    target_speed=to_float(
                        data.get(
                            "target_speed"
                        )
                    ),
                    target_consumption=to_float(
                        data.get(
                            "target_consumption"
                        )
                    ),
                    source_message="",
                )
            )

            # ------------------------------------------------
            # Build observation
            # ------------------------------------------------

            observation_data = {
                "reported_time": reported_time,

                "vessel_name": vessel_name,

                "position": clean_text(
                    data.get(
                        "position"
                    )
                ),

                "course": to_float(
                    data.get(
                        "course"
                    )
                ),

                "speed": to_float(
                    data.get(
                        "speed"
                    )
                ),

                "consumption": to_float(
                    data.get(
                        "consumption"
                    )
                ),

                "distance": to_float(
                    data.get(
                        "distance"
                    )
                ),

                "distance_run": to_float(
                    data.get(
                        "distance_run"
                    )
                ),

                "distance_to_go": to_float(
                    data.get(
                        "distance_to_go"
                    )
                ),

                "duration_days": to_float(
                    data.get(
                        "duration_days"
                    )
                ),

                "running_hours": to_float(
                    data.get(
                        "running_hours"
                    )
                ),

                "rpm": to_float(
                    data.get(
                        "rpm"
                    )
                ),

                "power_kw": to_float(
                    data.get(
                        "power_kw"
                    )
                ),

                "engine_power_kw": to_float(
                    data.get(
                        "engine_power_kw"
                    )
                ),

                "shaft_power_kw": to_float(
                    data.get(
                        "shaft_power_kw"
                    )
                ),

                "load_percent": to_float(
                    data.get(
                        "load_percent"
                    )
                ),

                "engine_load_percent": to_float(
                    data.get(
                        "engine_load_percent"
                    )
                ),

                "slip": to_float(
                    data.get(
                        "slip"
                    )
                ),

                "hsfo_consumption_mt": to_float(
                    data.get(
                        "hsfo_consumption_mt"
                    )
                ),

                "hsfo_rob": to_float(
                    data.get(
                        "hsfo_rob"
                    )
                ),

                "hsfo_rob_mt": to_float(
                    data.get(
                        "hsfo_rob_mt"
                    )
                ),

                "lsfo_consumption_mt": to_float(
                    data.get(
                        "lsfo_consumption_mt"
                    )
                ),

                "lsfo_rob": to_float(
                    data.get(
                        "lsfo_rob"
                    )
                ),

                "lsmgo_consumption_mt": to_float(
                    data.get(
                        "lsmgo_consumption_mt"
                    )
                ),

                "lsmgo_rob": to_float(
                    data.get(
                        "lsmgo_rob"
                    )
                ),

                "lsmgo_rob_mt": to_float(
                    data.get(
                        "lsmgo_rob_mt"
                    )
                ),

                "cylinder_oil_consumption_l": to_float(
                    data.get(
                        "cylinder_oil_consumption_l"
                    )
                ),

                "scoc": to_float(
                    data.get(
                        "scoc"
                    )
                ),

                "wind": clean_text(
                    data.get(
                        "wind"
                    )
                ),

                "swell": clean_text(
                    data.get(
                        "swell"
                    )
                ),

                "current": clean_text(
                    data.get(
                        "current"
                    )
                ),

                "eta": clean_text(
                    data.get(
                        "eta"
                    )
                ),

                "remarks": clean_text(
                    data.get(
                        "remarks"
                    )
                ),

                "source_file": getattr(
                    uploaded_file,
                    "name",
                    "",
                ),

                "source_message": "",
            }

            # ------------------------------------------------
            # Save observation
            # ------------------------------------------------

            observation, created = (
                save_observation(
                    leg,
                    observation_data,
                )
            )

            if created:
                imported += 1
            else:
                updated += 1

        # ----------------------------------------------------
        # Update leg summary
        # ----------------------------------------------------

        update_voyage_leg_summary(
            leg
        ) if "leg" in locals() else None

    return {
        "imported": imported,
        "updated": updated,
        "total": imported + updated,
    }