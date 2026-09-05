from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import load_workbook

from .forms import (
    VesselForm,
    InspectionForm,
    InspectionFindingForm,
    ChecklistItemForm,
    InspectionFindingRiskForm,
)

from .models import (
    Vessel,
    Inspection,
    ChecklistItem,
    InspectionFinding,
)


# =========================================================
# RIGHTSHIP RISK COLOUR DETECTION
# =========================================================

def get_excel_colour(cell):
    """
    Return the actual Excel colour information from a cell.

    Excel colours can be stored as:
        RGB
        ARGB
        indexed
        theme

    We primarily use RGB/ARGB because the RightShip workbook
    normally stores the inspection risk colours this way.
    """

    if cell is None:
        return None

    fill = cell.fill

    if fill is None:
        return None

    if fill.fill_type != "solid":
        return None

    color = fill.fgColor

    if color is None:
        return None

    # -----------------------------------------------------
    # RGB / ARGB
    # -----------------------------------------------------

    if color.type == "rgb":

        rgb = color.rgb

        if rgb:
            rgb = str(rgb).upper().replace("#", "")

            # Remove alpha channel if present
            if len(rgb) == 8:
                rgb = rgb[-6:]

            return rgb

    # -----------------------------------------------------
    # Indexed colours
    # -----------------------------------------------------

    if color.type == "indexed":

        indexed = color.indexed

        if indexed is not None:
            return f"INDEXED:{indexed}"

    # -----------------------------------------------------
    # Theme colours
    # -----------------------------------------------------

    if color.type == "theme":

        theme = color.theme

        if theme is not None:
            return f"THEME:{theme}"

    return None


def get_risk_from_excel_cell(cell):
    """
    Detect RightShip risk from the fill colour of the Ref No. cell.

    RightShip convention:

        Yellow  -> LOW
        Orange  -> MEDIUM
        Red     -> HIGH

    Supports RGB, indexed and theme colours.
    """

    fill = cell.fill

    if not fill:
        return None

    if fill.fill_type != "solid":
        return None

    # ---------------------------------------------------------
    # Try foreground colour
    # ---------------------------------------------------------

    color = fill.fgColor

    # ---------------------------------------------------------
    # RGB
    # ---------------------------------------------------------

    if color.type == "rgb" and color.rgb:

        rgb = color.rgb.upper().replace("#", "")

        # Remove alpha if present
        if len(rgb) == 8:
            rgb = rgb[2:]

        # Yellow
        yellow_values = {
            "FFFF00",
            "FFF200",
            "FFD966",
            "FFE699",
            "FFFF00",
        }

        # Orange
        orange_values = {
            "FFC000",
            "F4B183",
            "ED7D31",
            "FF9900",
            "F39C12",
        }

        # Red
        red_values = {
            "FF0000",
            "C00000",
            "FF5050",
            "E74C3C",
            "D9534F",
        }

        if rgb in yellow_values:
            return "LOW"

        if rgb in orange_values:
            return "MEDIUM"

        if rgb in red_values:
            return "HIGH"

    # ---------------------------------------------------------
    # Indexed colours
    # ---------------------------------------------------------

    if color.type == "indexed":

        indexed = color.indexed

        # Common Excel indexed colours
        if indexed in {6, 13}:
            return "LOW"

        if indexed in {45, 52}:
            return "MEDIUM"

        if indexed in {10, 9}:
            return "HIGH"

    return None

# =========================================================
# DEBUG HELPER
# =========================================================

def inspect_excel_colour(cell):
    """
    Useful while testing the workbook.

    Returns a dictionary describing how openpyxl sees
    the cell colour.
    """

    fill = cell.fill
    color = fill.fgColor if fill else None

    return {
        "coordinate": cell.coordinate,
        "fill_type": fill.fill_type if fill else None,
        "color_type": color.type if color else None,
        "rgb": color.rgb if color else None,
        "indexed": color.indexed if color else None,
        "theme": color.theme if color else None,
    }


# =========================================================
# VESSEL LIST
# =========================================================

@login_required
def vessel_list(request):

    vessels = (
        Vessel.objects
        .all()
        .order_by("vessel_name")
    )

    return render(
        request,
        "inspections/vessel_list.html",
        {
            "vessels": vessels,
        },
    )


# =========================================================
# CREATE VESSEL
# =========================================================

@login_required
def vessel_create(request):

    if request.method == "POST":

        form = VesselForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Vessel created successfully.",
            )

            return redirect("vessel_list")

    else:

        form = VesselForm()

    return render(
        request,
        "inspections/vessel_form.html",
        {
            "form": form,
        },
    )


# =========================================================
# EDIT VESSEL
# =========================================================

@login_required
def vessel_edit(request, pk):

    vessel = get_object_or_404(
        Vessel,
        pk=pk,
    )

    if request.method == "POST":

        form = VesselForm(
            request.POST,
            instance=vessel,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Vessel updated successfully.",
            )

            return redirect("vessel_list")

    else:

        form = VesselForm(
            instance=vessel,
        )

    return render(
        request,
        "inspections/vessel_form.html",
        {
            "form": form,
            "vessel": vessel,
        },
    )


# =========================================================
# DELETE VESSEL
# =========================================================

@login_required
def vessel_delete(request, pk):

    vessel = get_object_or_404(
        Vessel,
        pk=pk,
    )

    if request.method == "POST":

        vessel.delete()

        messages.success(
            request,
            "Vessel deleted successfully.",
        )

        return redirect("vessel_list")

    return render(
        request,
        "inspections/vessel_confirm_delete.html",
        {
            "vessel": vessel,
        },
    )


    # =========================================================
    # INSPECTION LIST
    # =========================================================

@login_required
def inspection_list(request):

    inspections = (
        Inspection.objects
        .select_related("vessel")
        .prefetch_related("findings")
        .order_by("-inspection_date", "-id")
    )

    # =====================================================
    # KPI COUNTS
    # =====================================================

    total_inspections = inspections.count()

    open_inspections = inspections.filter(
        status="OPEN"
    ).count()

    completed_inspections = inspections.filter(
        status="COMPLETED"
    ).count()

    total_vessels = Vessel.objects.count()

    # =====================================================
    # RISK COUNTS
    # =====================================================

    high_risk = InspectionFinding.objects.filter(
        risk_level="HIGH"
    ).count()

    medium_risk = InspectionFinding.objects.filter(
        risk_level="MEDIUM"
    ).count()

    low_risk = InspectionFinding.objects.filter(
        risk_level="LOW"
    ).count()

    # =====================================================
    # RENDER
    # =====================================================

    return render(
        request,
        "inspections/inspection_list.html",
        {
            "inspections": inspections,

            "total_inspections": total_inspections,
            "open_inspections": open_inspections,
            "completed_inspections": completed_inspections,
            "total_vessels": total_vessels,

            "high_risk": high_risk,
            "medium_risk": medium_risk,
            "low_risk": low_risk,
        },
    )


# =========================================================
# CREATE INSPECTION
# =========================================================

@login_required
def inspection_create(request):

    if request.method == "POST":

        form = InspectionForm(request.POST)

        if form.is_valid():

            inspection = form.save(
                commit=False
            )

            year = timezone.now().year

            last = (
                Inspection.objects
                .filter(
                    inspection_no__startswith=f"INS-{year}-"
                )
                .order_by("-id")
                .first()
            )

            if last:

                last_no = int(
                    last.inspection_no
                    .split("-")[-1]
                )

                next_no = last_no + 1

            else:

                next_no = 1

            inspection.inspection_no = (
                f"INS-{year}-{next_no:05d}"
            )

            inspection.save()

            messages.success(
                request,
                "Inspection created successfully.",
            )

            return redirect(
                "inspection_list"
            )

    else:

        form = InspectionForm()

    return render(
        request,
        "inspections/inspection_form.html",
        {
            "form": form,
        },
    )


# =========================================================
# INSPECTION FINDINGS
# =========================================================

@login_required
def inspection_findings(request, pk):

    inspection = get_object_or_404(
        Inspection.objects.select_related("vessel"),
        pk=pk,
    )

    findings = (
        inspection.findings
        .select_related("checklist_item")
        .order_by(
            "checklist_item__ref_no"
        )
    )

    # =====================================================
    # RISK COUNTS
    # =====================================================

    total_findings = findings.count()

    high_risk = findings.filter(
        risk_level="HIGH"
    ).count()

    medium_risk = findings.filter(
        risk_level="MEDIUM"
    ).count()

    low_risk = findings.filter(
        risk_level="LOW"
    ).count()

    # =====================================================
    # CONTEXT
    # =====================================================

    context = {
        "inspection": inspection,
        "findings": findings,

        "total_findings": total_findings,
        "high_risk": high_risk,
        "medium_risk": medium_risk,
        "low_risk": low_risk,
    }

    return render(
        request,
        "inspections/inspection_findings.html",
        context,
    )


# =========================================================
# CREATE FINDING
# =========================================================

@login_required
def finding_create(
    request,
    inspection_id,
):

    inspection = get_object_or_404(
        Inspection,
        pk=inspection_id,
    )

    used_items = (
        InspectionFinding.objects
        .filter(
            inspection=inspection,
        )
        .values_list(
            "checklist_item_id",
            flat=True,
        )
    )

    if request.method == "POST":

        form = InspectionFindingForm(
            request.POST,
            inspection=inspection,
        )

        form.fields[
            "checklist_item"
        ].queryset = (
            ChecklistItem.objects
            .exclude(
                id__in=used_items,
            )
            .order_by("ref_no")
        )

        if form.is_valid():

            finding = form.save(
                commit=False
            )

            finding.inspection = inspection

            finding.save()

            messages.success(
                request,
                "Finding added successfully.",
            )

            return redirect(
                "finding_create",
                inspection_id=inspection.id,
            )

    else:

        form = InspectionFindingForm(
            inspection=inspection,
        )

        form.fields[
            "checklist_item"
        ].queryset = (
            ChecklistItem.objects
            .exclude(
                id__in=used_items,
            )
            .order_by("ref_no")
        )

    findings = (
        InspectionFinding.objects
        .filter(
            inspection=inspection,
        )
        .select_related(
            "checklist_item",
        )
        .order_by(
            "checklist_item__ref_no",
        )
    )

    return render(
        request,
        "inspections/finding_form.html",
        {
            "inspection": inspection,
            "form": form,
            "findings": findings,
        },
    )


# =========================================================
# EDIT FINDING RISK
# =========================================================

@login_required
def finding_edit_risk(request, pk):

    finding = get_object_or_404(
        InspectionFinding,
        pk=pk,
    )

    if request.method == "POST":

        form = InspectionFindingRiskForm(
            request.POST,
            instance=finding,
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Risk level updated successfully.",
            )

            return redirect(
                "finding_create",
                inspection_id=finding.inspection.id,
            )

    else:

        form = InspectionFindingRiskForm(
            instance=finding,
        )

    return render(
        request,
        "inspections/finding_edit_risk.html",
        {
            "finding": finding,
            "form": form,
        },
    )

# =========================================================
# DELETE FINDING
# =========================================================

@login_required
def finding_delete(request, pk):

    finding = get_object_or_404(
        InspectionFinding,
        pk=pk,
    )

    inspection_id = finding.inspection.id

    if request.method == "POST":

        finding.delete()

        messages.success(
            request,
            "Finding deleted successfully.",
        )

        return redirect(
            "inspection_findings",
            pk=inspection_id,
        )

    return redirect(
        "inspection_findings",
        pk=inspection_id,
    )
# =========================================================
# CHECKLIST ITEM CREATE
# =========================================================

@login_required
def checklistitem_create(request):

    if request.method == "POST":

        form = ChecklistItemForm(
            request.POST
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Checklist item created successfully.",
            )

            return redirect(
                "checklistitem_list"
            )

    else:

        form = ChecklistItemForm()

    return render(
        request,
        "inspections/checklistitem_form.html",
        {
            "form": form,
        },
    )


# =========================================================
# CHECKLIST ITEM LIST
# =========================================================

@login_required
def checklistitem_list(request):

    items = (
        ChecklistItem.objects
        .all()
        .order_by("ref_no")
    )

    return render(
        request,
        "inspections/checklistitem_list.html",
        {
            "items": items,
        },
    )


# =========================================================
# RIGHTSHIP EXCEL IMPORT
# =========================================================

@login_required
@transaction.atomic
def inspection_import(request):

    sheet_data = []

    if request.method != "POST":

        return render(
            request,
            "inspections/inspection_import.html",
            {
                "sheet_data": sheet_data,
            },
        )

    excel_file = request.FILES.get(
        "excel_file"
    )

    if not excel_file:

        messages.error(
            request,
            "Please select an Excel workbook.",
        )

        return render(
            request,
            "inspections/inspection_import.html",
            {
                "sheet_data": sheet_data,
            },
        )

    try:

        workbook = load_workbook(
            excel_file,
            data_only=True,
        )

    except Exception as exc:

        messages.error(
            request,
            f"Could not read Excel workbook: {exc}",
        )

        return render(
            request,
            "inspections/inspection_import.html",
            {
                "sheet_data": sheet_data,
            },
        )

    skip_sheets = {
        "Index",
        "GRAPH",
        "Sheet1",
    }

    total_created = 0
    total_updated = 0
    total_skipped = 0

    total_low = 0
    total_medium = 0
    total_high = 0
    total_unclassified = 0

    # =====================================================
    # EACH VESSEL SHEET
    # =====================================================

    for ws in workbook.worksheets:

        if ws.title in skip_sheets:
            continue

        vessel_name = str(
            ws.title
        ).strip()

        if not vessel_name:
            continue

        print(
        "IMPORT DEBUG:",
        ws.title,
        "| D3:", ws["D3"].value,
        "| D4:", ws["D4"].value,
        "| D5:", ws["D5"].value,
        "| D6:", ws["D6"].value,
        "| D7:", ws["D7"].value,
    )

        # =================================================
        # HEADER
        # =================================================

        port = ws["D3"].value

        inspection_date = ws["D4"].value

        inspector = ws["D5"].value

        findings = ws["D6"].value

        validity_text = ws["D7"].value

        # =================================================
        # VALIDITY
        # =================================================

        validity = 6

        if validity_text:

            try:

                validity = int(
                    str(
                        validity_text
                    ).split()[0]
                )

            except (
                ValueError,
                TypeError,
            ):

                validity = 6

        # =================================================
        # VESSEL
        # =================================================

        vessel, vessel_created = (
            Vessel.objects.get_or_create(
                vessel_name=vessel_name,
            )
        )

        # =================================================
        # INSPECTION
        # =================================================

        inspection = (
            Inspection.objects
            .filter(
                vessel=vessel,
                inspection_date=inspection_date,
            )
            .first()
        )

        inspection_created = False

        if inspection is None:

            if hasattr(
                inspection_date,
                "year",
            ):

                year = (
                    inspection_date.year
                )

            else:

                year = timezone.now().year

            last = (
                Inspection.objects
                .filter(
                    inspection_no__startswith=(
                        f"INS-{year}-"
                    )
                )
                .order_by("-id")
                .first()
            )

            if last:

                next_no = (
                    int(
                        last.inspection_no
                        .split("-")[-1]
                    )
                    + 1
                )

            else:

                next_no = 1

            inspection = (
                Inspection.objects.create(

                    inspection_no=(
                        f"INS-{year}-{next_no:05d}"
                    ),

                    vessel=vessel,

                    port=port or "",

                    inspection_date=(
                        inspection_date
                    ),

                    inspector=(
                        inspector or ""
                    ),

                    validity_months=validity,

                    status="OPEN",

                    remarks=(
                        "Imported from RightShip Excel"
                    ),
                )
            )

            inspection_created = True

        else:

            inspection.port = port or ""

            inspection.inspector = (
                inspector or ""
            )

            inspection.validity_months = (
                validity
            )

            inspection.save(
                update_fields=[
                    "port",
                    "inspector",
                    "validity_months",
                ]
            )

        # =================================================
        # COUNTERS
        # =================================================

        findings_created = 0
        findings_updated = 0
        findings_skipped = 0

        low_count = 0
        medium_count = 0
        high_count = 0
        unclassified_count = 0

        # =================================================
        # FINDINGS
        # =================================================

        for row in range(
            9,
            ws.max_row + 1,
        ):

            ref_cell = ws[f"B{row}"]

            ref_no = ref_cell.value

            inspected_item = (
                ws[f"C{row}"].value
            )

            finding_description = (
                ws[f"D{row}"].value
            )

            # Skip completely empty rows
            if ref_no is None:
                continue

            ref_no = str(
                ref_no
            ).strip()

            if not ref_no:
                continue

            inspected_item = (
                str(
                    inspected_item
                ).strip()
                if inspected_item is not None
                else ""
            )

            finding_description = (
                str(
                    finding_description
                ).strip()
                if finding_description is not None
                else ""
            )

            # =================================================
            # RISK FROM EXCEL COLOUR
            # =================================================

            risk_level = (
                get_risk_from_excel_cell(
                    ref_cell
                )
            )

            # =================================================
            # CHECKLIST ITEM
            # =================================================

            checklist_item, _ = (
                ChecklistItem.objects
                .get_or_create(

                    ref_no=ref_no,

                    defaults={
                        "inspected_item":
                            inspected_item,
                    },
                )
            )

            if (
                inspected_item
                and
                checklist_item.inspected_item
                != inspected_item
            ):

                checklist_item.inspected_item = (
                    inspected_item
                )

                checklist_item.save(
                    update_fields=[
                        "inspected_item"
                    ]
                )

            # =================================================
            # EXISTING FINDING
            # =================================================

            finding = (
                InspectionFinding.objects
                .filter(
                    inspection=inspection,
                    checklist_item=checklist_item,
                )
                .first()
            )

            # =================================================
            # NO RECOGNISED COLOUR
            # =================================================

            if risk_level is None:

                unclassified_count += 1
                total_unclassified += 1

                if finding:

                    changed = False

                    if (
                        finding.finding_description
                        != finding_description
                    ):

                        finding.finding_description = (
                            finding_description
                        )

                        changed = True

                    if changed:

                        finding.save(
                            update_fields=[
                                "finding_description"
                            ]
                        )

                        findings_updated += 1
                        total_updated += 1

                continue

            # =================================================
            # CREATE
            # =================================================

            if finding is None:

                InspectionFinding.objects.create(

                    inspection=inspection,

                    checklist_item=(
                        checklist_item
                    ),

                    finding_description=(
                        finding_description
                    ),

                    risk_level=risk_level,
                )

                findings_created += 1
                total_created += 1

            # =================================================
            # UPDATE
            # =================================================

            else:

                changed = False

                if (
                    finding.finding_description
                    != finding_description
                ):

                    finding.finding_description = (
                        finding_description
                    )

                    changed = True

                if (
                    finding.risk_level
                    != risk_level
                ):

                    finding.risk_level = (
                        risk_level
                    )

                    changed = True

                if changed:

                    finding.save()

                    findings_updated += 1
                    total_updated += 1

                else:

                    findings_skipped += 1
                    total_skipped += 1

            # =================================================
            # RISK COUNTS
            # =================================================

            if risk_level == "LOW":

                low_count += 1
                total_low += 1

            elif risk_level == "MEDIUM":

                medium_count += 1
                total_medium += 1

            elif risk_level == "HIGH":

                high_count += 1
                total_high += 1

        # =================================================
        # SHEET SUMMARY
        # =================================================

        sheet_data.append({

            "inspection_no":
                inspection.inspection_no,

            "vessel":
                vessel_name,

            "port":
                port,

            "inspection_date":
                inspection_date,

            "inspector":
                inspector,

            "findings":
                findings,

            "validity":
                validity,

            "vessel_created":
                vessel_created,

            "inspection_created":
                inspection_created,

            "findings_created":
                findings_created,

            "findings_updated":
                findings_updated,

            "findings_skipped":
                findings_skipped,

            "low_count":
                low_count,

            "medium_count":
                medium_count,

            "high_count":
                high_count,

            "unclassified_count":
                unclassified_count,
        })

    # =====================================================
    # MESSAGES
    # =====================================================

    messages.success(
        request,
        (
            f"Import completed: "
            f"{len(sheet_data)} inspection sheet(s), "
            f"{total_created} new finding(s), "
            f"{total_updated} updated finding(s)."
        ),
    )

    messages.info(
        request,
        (
            f"Risk distribution: "
            f"{total_high} High / "
            f"{total_medium} Medium / "
            f"{total_low} Low."
        ),
    )

    if total_unclassified:

        messages.warning(
            request,
            (
                f"{total_unclassified} row(s) "
                f"could not be classified from "
                f"the Excel colour."
            ),
        )

    return render(
        request,
        "inspections/inspection_import.html",
        {
            "sheet_data": sheet_data,
        },
    )

# =========================================================
# RIGHTSHIP PERFORMANCE DASHBOARD
# =========================================================

@login_required
def rightship_dashboard(request):

    # =====================================================
    # BASIC COUNTS
    # =====================================================

    total_vessels = Vessel.objects.count()

    total_inspections = Inspection.objects.count()

    total_findings = InspectionFinding.objects.count()

    # =====================================================
    # RISK COUNTS
    # =====================================================

    high_risk = InspectionFinding.objects.filter(
        risk_level="HIGH"
    ).count()

    medium_risk = InspectionFinding.objects.filter(
        risk_level="MEDIUM"
    ).count()

    low_risk = InspectionFinding.objects.filter(
        risk_level="LOW"
    ).count()

    # =====================================================
    # INSPECTION DATA
    # =====================================================

    inspections = (
        Inspection.objects
        .select_related("vessel")
        .order_by("inspection_date", "id")
    )

    inspection_rows = []

    for inspection in inspections:

        findings = InspectionFinding.objects.filter(
            inspection=inspection
        )

        high = findings.filter(
            risk_level="HIGH"
        ).count()

        medium = findings.filter(
            risk_level="MEDIUM"
        ).count()

        low = findings.filter(
            risk_level="LOW"
        ).count()

        total = findings.count()

        # -------------------------------------------------
        # SEVERITY
        # -------------------------------------------------

        if high > 0:
            severity = 3

        elif medium > 0:
            severity = 2

        elif low > 0:
            severity = 1

        else:
            severity = 0

        inspection_rows.append(
            {
                "inspection": inspection,
                "vessel": inspection.vessel.vessel_name,
                "date": inspection.inspection_date,
                "high": high,
                "medium": medium,
                "low": low,
                "total": total,
                "severity": severity,
                "validity": inspection.validity_months,
            }
        )

    # =====================================================
    # CHART DATA
    # =====================================================

    chart_labels = []
    high_data = []
    medium_data = []
    low_data = []
    severity_data = []
    validity_data = []

    for row in inspection_rows:

        inspection_date = row["date"]

        if inspection_date:
            date_text = inspection_date.strftime(
                "%d-%b-%y"
            )
        else:
            date_text = ""

        chart_labels.append(
            f"{row['vessel']} - {date_text}"
        )

        high_data.append(
            row["high"]
        )

        medium_data.append(
            row["medium"]
        )

        low_data.append(
            row["low"]
        )

        severity_data.append(
            row["severity"]
        )

        validity_data.append(
            row["validity"] or 0
        )

    # =====================================================
    # CONTEXT
    # =====================================================

    context = {
        "total_vessels": total_vessels,

        "total_inspections": total_inspections,

        "total_findings": total_findings,

        "high_risk": high_risk,

        "medium_risk": medium_risk,

        "low_risk": low_risk,

        "inspections_register": inspection_rows,

        "chart_labels": chart_labels,

        "high_data": high_data,

        "medium_data": medium_data,

        "low_data": low_data,

        "severity_data": severity_data,

        "validity_data": validity_data,
    }

    return render(
        request,
        "inspections/dashboard.html",
        context,
    )